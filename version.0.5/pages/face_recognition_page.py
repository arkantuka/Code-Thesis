import os
import customtkinter as ctk
import cv2
import datetime

import openpyxl
import pages.choose_course_face_recognition as choose_f_rec
from PIL import Image

class FaceRecognition:
    
    # Back Button Function
    def back(window, frame):
        frame.destroy()
        cam1 = cv2.VideoCapture(0+cv2.CAP_DSHOW)
        cam1.release()
        choose_f_rec.ChooseCourseFaceRecognition(window)
        
    def writeAttendance(id, name, course_name):
        parent_dir = 'version.0.5/time_attendance'
        file_path = os.path.join(parent_dir, course_name+'.xlsx')
        if not os.path.exists(file_path):
            workbook = openpyxl.Workbook()
            workbook.create_sheet(course_name)
            sheetname = workbook.sheetnames
            sheet = workbook[sheetname[0]]
            sheet.append(['Student ID', 'Name','Date', 'Time'])
            workbook.save(file_path)
        workbook = openpyxl.load_workbook(file_path)
        sheetname = workbook.sheetnames
        sheet = workbook[sheetname[0]]
        sheet.append([id, name,datetime.datetime.now().strftime("%d/%m/%Y"), datetime.datetime.now().strftime("%H:%M:%S")])
        workbook.save(file_path)
        workbook.close()
        
    def create_camera_frame(frame, course_name, studentIDs, studentNames):
        # Create Camera Frame
        cam_frame = ctk.CTkFrame(master=frame)
        cam_frame.pack(pady=(50, 10))
        
        already_detected = []
        
        directory = os.path.dirname(__file__)
        weights = os.path.join(directory, "face_detection_yunet_2023mar.onnx")
        face_cascade_yunet = cv2.FaceDetectorYN_create(weights, "", (0, 0))
        
        width = 1080
        height = 800
        cam = cv2.VideoCapture(0+cv2.CAP_DSHOW)
        cam.set(cv2.CAP_PROP_FRAME_WIDTH, width)
        cam.set(cv2.CAP_PROP_FRAME_HEIGHT, height)
        lmain = ctk.CTkLabel(cam_frame, text="")
        lmain.pack()
        face_recognizer = cv2.face.LBPHFaceRecognizer_create()
        model_path = 'version.0.5/training_models/'+course_name+'.yml'
        face_recognizer.read(model_path)
        
        def show_frame():
            _, frame = cam.read()
            cv2image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGBA)
            gray_img = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            
            h, w, _ = frame.shape
            faces_yn = face_cascade_yunet.setInputSize((w, h))
            _, faces_yn = face_cascade_yunet.detect(frame)
            faces_yn = faces_yn if faces_yn is not None else []
            
            for face in faces_yn:
                
                box = list(map(int, face[:4]))
                color = (255, 0, 0, 255)
                thickness = 2
                cv2.rectangle(cv2image, box, color, thickness, lineType=cv2.LINE_AA)
                try:
                    serial, confidence = face_recognizer.predict(gray_img[box[1]:box[1]+box[3], box[0]:box[0]+box[2]])
                    print(serial, confidence)
                except:
                    serial, confidence = 0, 0
                position = (box[0], box[1] - 10)
                font = cv2.FONT_HERSHEY_SIMPLEX
                scale = 0.5
                thickness = 2
                color_red = (255, 0, 0, 255)
                color_green = (0, 255, 0, 255)
                if confidence > 70:
                    index = studentIDs.index(str(serial))
                    print(index)
                    cv2.putText(cv2image, str(serial)+" "+"{:.2f}".format(confidence), position, font, scale, color_green, thickness, lineType=cv2.LINE_AA)
                    cv2.rectangle(cv2image, box, color_green, thickness, lineType=cv2.LINE_AA)
                    if str(serial) not in already_detected:
                        FaceRecognition.writeAttendance(studentIDs[index], studentNames[index], course_name)
                        already_detected.append(str(serial))
                else:
                    name = "Unknown"
                    cv2.putText(cv2image, name, position, font, scale, color_red, thickness, lineType=cv2.LINE_AA)
            
            img = Image.fromarray(cv2image)
            imgctk = ctk.CTkImage(light_image=img, dark_image=img, size=img.size)
            lmain.configure(image=imgctk)
            lmain.after(100, show_frame)
        show_frame()
        
        
    def __init__(self, window, file_path, studentIDs, studentNames):
        # Create Master Frame
        master_frame = ctk.CTkFrame(master=window)
        master_frame.pack(fill="both", expand=True)
        master_frame.grid_rowconfigure(0, weight=1)
        master_frame.grid_columnconfigure(0, weight=1)
        course_name = str(file_path)[31:-5]
        
        # Create Camera Frame
        FaceRecognition.create_camera_frame(master_frame, course_name, studentIDs, studentNames)
        
        # Back without save Button
        back_button = ctk.CTkButton(master=master_frame, fg_color='red',
                                    width=220, height=50,
                                    text="Back",
                                    font=("Leelawadee", 25),
                                    command=lambda: FaceRecognition.back(window, master_frame))
        back_button.pack(side="bottom", pady=20)
