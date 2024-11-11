import os
import customtkinter as ctk
import openpyxl
import pages.menu_manage_course as menu_mng_crs
from tkinter import ttk

class CourseDetail:
    # Back Button Function
    def back(window, frame):
        frame.destroy()
        menu_mng_crs.ManageCourseMenu(window)
        
    # Back and save Button Function
    def backAndSave(window, frame, workbook, table, file_path, dict_of_data):
        sheetname = workbook.sheetnames
        sheet = workbook[sheetname[0]]
        sheet.delete_rows(1, sheet.max_row)
        sheet.append(list(dict_of_data.keys()))
        sheet.append(list(dict_of_data.values()))
        header = ['Student ID', 'Name', 'Face Status']
        sheet.append(header)
        for row_id in table.get_children():
            row = table.item(row_id)['values'][:3]
            sheet.append(row)            
        workbook.save(file_path)
        workbook.close()
        frame.destroy()
        menu_mng_crs.ManageCourseMenu(window)
    
    def find_index_of_ids_duplicate(ids_in_database, ids_in_course):
        idx = []
        for id in ids_in_course:
            if id in ids_in_database:
                idx.append(ids_in_course.index(id)+3)
        return idx
    
    def update_face_status(ids):
        collection_ids = []
        year_path = os.listdir("version.0.5/images")
        for year in year_path:
            faceID_path = os.listdir(f"version.0.5/images/{year}")
            for id in faceID_path:
                collection_ids.append(id)
        print(ids)
        return CourseDetail.find_index_of_ids_duplicate(collection_ids, ids)
    
    # Create Button
    def createButton(frame, text, command):
        button = ctk.CTkButton(master=frame, fg_color='gray',
                                width=200, height=50,
                                text=text,
                                font=("Leelawadee", 25),
                                command=command)
        return button
    
    # Create Entry
    def createEntry(frame, wid, hold_text):
        entry = ctk.CTkEntry(frame, width=wid,
                             font=('Leelawadee', 20), text_color='black',
                             placeholder_text=hold_text,
                             fg_color="#f0f0f0")
        return entry
    
    # Select record Function
    def select_record(table, id_entry, name_entry):
        # Clear entry boxs
        id_entry.delete(0, 'end')
        name_entry.delete(0, 'end')
        # Grab record
        selected = table.focus()
        values = table.item(selected, 'values')
        # Output to boxs
        id_entry.insert(0, values[0])
        name_entry.insert(0, values[1])
        
    # Insert Data Function
    def add(table, id_entry, name_entry):
        if id_entry.get() and name_entry.get() != '':
            table.insert(parent='', index='end', text="",
                        values=(id_entry.get(), name_entry.get(), 'None'))
            # Clear the boxes
            id_entry.delete(0, 'end')
            name_entry.delete(0, 'end')
        else:
            pass
        
    # Update Record Function
    def edit(table, id_entry, name_entry):
        if id_entry.get() and name_entry.get() != '':
            # Grab record number
            selected = table.focus()
            face_status = table.item(selected, 'values')[2]
            # Save new data
            table.item(selected, text="", values=(id_entry.get(), name_entry.get(), face_status))
        else:
            pass
        
    # Delete Record Function
    def delete(table):
        if table.selection() != () :   
            for item in table.selection():
                table.delete(item)
        else:
            pass
        
    def create_frame(window, master_frame, file_path):
        # Load Excel Data
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        loadToList = list(sheet.values)
        list_of_data = []
        for row in loadToList:
            list_of_data.append(list(row))
        dict_of_data = {list_of_data[0][i]: list_of_data[1][i] for i in range(len(list_of_data[0]))}
        list_of_data[2][2] = 'Face Status'
        allIDs = [i[0] for i in list_of_data[3:]]
        duplicate_index = CourseDetail.update_face_status(allIDs)
        for i in duplicate_index:
            list_of_data[i][2] = 'Collected'
        
        # Main Label
        main_label = ctk.CTkLabel(master=master_frame,
                                  text="Course Details",
                                  font=("Leelawadee", 35, "bold"),
                                  padx=100, pady=10)
        main_label.pack(pady=(15, 0))
        
        # Create Main Frame
        main_frame = ctk.CTkFrame(master=master_frame)
        main_frame.grid_columnconfigure(0, weight=1)
        main_frame.grid_columnconfigure(1, weight=1)
        main_frame.pack()
        
        # Create Frame Right
        frame_right = ctk.CTkFrame(master=main_frame)
        frame_right.grid(row=0, column=1, padx=(10, 0), pady=10)
        
        # Create Label
        student_label = ctk.CTkLabel(master=frame_right,
                                     text="Student Information",
                                     font=("Leelawadee", 20),
                                     padx=10, pady=10)
        student_label.pack()
        
        # Create course detail frame
        course_detail_frame = ctk.CTkFrame(master=frame_right)
        course_detail_frame.pack(pady=10)
        
        # Create table frame
        table_frame = ctk.CTkFrame(master=course_detail_frame)
        table_frame.pack()
        
        # Create Scrollbar and Table
        tree_scrollbar = ttk.Scrollbar(table_frame, orient="vertical")
        tree_scrollbar.pack(side="right", fill="y")
        table = ttk.Treeview(table_frame, height=20, show="headings",
                                     yscrollcommand=tree_scrollbar.set)
        table.pack(fill="both", expand=True)
        tree_scrollbar.config(command=table.yview)
        
        # Configure Table
        style = ttk.Style()
        style.configure("Treeview.Heading", font=("Leelawadee", 12, "bold"))
        style.configure("Treeview", font=("Leelawadee", 15), rowheight=20)

        # Insert Table
        table['columns'] = list_of_data[2][:3]
        table.column('#0', width=0, stretch='YES')
        table.column('Student ID', anchor='center', width=130)
        table.column('Name', width=260)
        table.column('Face Status', anchor='center', width=100)
        table.heading(str(list_of_data[2][0]), text=str(list_of_data[2][0]))
        table.heading(str(list_of_data[2][1]), text=str(list_of_data[2][1]))
        table.heading(str(list_of_data[2][2]), text=str(list_of_data[2][2]))
        for value in list_of_data[3:]:
            table.insert("", "end", values=value)
            
        # Create Student Entry Frame
        entry_frame = ctk.CTkFrame(master=frame_right)
        entry_frame.pack()
        
        # Create Entry for show selection data
        style.configure("Entry", font=('Leelawadee', 20, 'bold'))
        id_label = ctk.CTkLabel(entry_frame, text='Student ID')
        id_label.grid(row=0, column=0)
        id_entry = CourseDetail.createEntry(entry_frame, 150, 'Student ID')
        id_entry.grid(row=1, column=0)
        name_label = ctk.CTkLabel(entry_frame, text='Full Name')
        name_label.grid(row=0, column=1)
        name_entry = CourseDetail.createEntry(entry_frame, 350, 'Full Name')
        name_entry.grid(row=1, column=1)
        
        # Create Course Description Label
        course_description_label = ctk.CTkLabel(master=frame_right,
                                                text="Course Description",
                                                font=("Leelawadee", 20),
                                                padx=100, pady=10)
        course_description_label.pack()
        
        # Create Course Description Frame
        course_description_frame = ctk.CTkFrame(master=frame_right)
        course_description_frame.pack()
        
        # Create Information Frame
        courseID_label = ctk.CTkLabel(master=course_description_frame,
                                     text="ID: "+dict_of_data['Course ID'],
                                     font=("Leelawadee", 20))
        courseID_label.grid(row=0, column=0, padx= 20)
        courseName_label = ctk.CTkLabel(master=course_description_frame,
                                       text="Course Name: "+dict_of_data['Course Name'],
                                       font=("Leelawadee", 20))
        courseName_label.grid(row=0, column=1, padx= 20)
        courseTerm_label = ctk.CTkLabel(master=course_description_frame,
                                       text="Term: "+dict_of_data['Term'],
                                       font=("Leelawadee", 20))
        courseTerm_label.grid(row=1, column=0)
        courseYear_label = ctk.CTkLabel(master=course_description_frame,
                                       text="Year: "+dict_of_data['Year'],
                                       font=("Leelawadee", 20))
        courseYear_label.grid(row=1, column=1)
        course_midterm_date_label = ctk.CTkLabel(master=course_description_frame,
                                                text="Midterm Date: "+dict_of_data['Midterm Exam'],
                                                font=("Leelawadee", 20))
        course_midterm_date_label.grid(row=2, column=0, padx= 20)
        course_midterm_time_label = ctk.CTkLabel(master=course_description_frame,
                                                text="Midterm Time: "+dict_of_data['Midterm Start Time'],
                                                font=("Leelawadee", 20))
        course_midterm_time_label.grid(row=3, column=0)
        course_final_date_label = ctk.CTkLabel(master=course_description_frame,
                                               text="Final Date: "+dict_of_data['Final Exam'],
                                               font=("Leelawadee", 20))
        course_final_date_label.grid(row=2, column=1, padx= 20)
        course_final_time_label = ctk.CTkLabel(master=course_description_frame,
                                               text="Final Time: "+dict_of_data['Final Start Time'],
                                               font=("Leelawadee", 20))
        course_final_time_label.grid(row=3, column=1)        
        
        # Create Frame Left
        frame_left = ctk.CTkFrame(master=main_frame)
        frame_left.grid(row=0, column=0, padx=(0, 10), pady=10)
        
        # Create Buttons Frame
        buttons_frame = ctk.CTkFrame(master=frame_left)
        buttons_frame.pack(pady=10)
        
        # Create Add Button
        add_button = CourseDetail.createButton(buttons_frame, 'Add', lambda: CourseDetail.add(table, id_entry, name_entry))
        add_button.grid(row=0, column=0, padx=20, pady=10)
        
        # Create Edit Button
        edit_button = CourseDetail.createButton(buttons_frame, 'Edit', lambda: CourseDetail.edit(table, id_entry, name_entry))
        edit_button.grid(row=1, column=0, padx=20, pady=10)
        
        # Create Delete Button
        delete_button = CourseDetail.createButton(buttons_frame, 'Delete', lambda: CourseDetail.delete(table))
        delete_button.grid(row=2, column=0, padx=20, pady=10)
        
        # bind
        table.bind('<Delete>', lambda e: CourseDetail.delete(table))
        table.bind('<ButtonRelease-1>', lambda e: CourseDetail.select_record(table, id_entry, name_entry))
        
        # Back and save Button
        back_n_save_button = ctk.CTkButton(master=frame_left, fg_color='green',
                                    width=220, height=50,
                                    text="Back and Save",
                                    font=("Leelawadee", 25),
                                    command=lambda: CourseDetail.backAndSave(window, master_frame, workbook, table, file_path, dict_of_data))
        back_n_save_button.pack(side="bottom", pady=(30,30))
        
        # Back Button
        back_button = ctk.CTkButton(master=frame_left, fg_color='red',
                                    width=220, height=50,
                                    text="Back without save",
                                    font=("Leelawadee", 25),
                                    command=lambda: CourseDetail.back(window, master_frame))
        back_button.pack(side="bottom", pady=(100, 0))
    
    def __init__(self, window, file_path):
        # Create Master Frame
        master_frame = ctk.CTkFrame(master=window)
        master_frame.pack(fill="both", expand=True)
        master_frame.grid_rowconfigure(0, weight=1)
        master_frame.grid_columnconfigure(0, weight=1)
        
        CourseDetail.create_frame(window, master_frame, file_path)