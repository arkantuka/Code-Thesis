import threading
import customtkinter as ctk
import openpyxl
import cv2
import pages.student_data_page as std_page
from PIL import Image
from tkinter import ttk

# Create Camera Thread
class CameraThread(threading.Thread):
    def __init__(self, camera_id, camera_frame):
        threading.Thread.__init__(self)
        self.camera_id = camera_id
        self.camera_frame = camera_frame
        
    def run(self):
        ShowStudentDataPage.capture_video(self.camera_id, self.camera_frame)


class ShowStudentDataPage:
    
    # Back Button Function
    def back(window, frame):
        
        frame.destroy()
        std_page.StudentDataPage(window)
    
    # Back and save Button Function
    def backAndSave(window, frame, workbook, table, file_path):
        sheetname = workbook.sheetnames
        sheet = workbook[sheetname[0]]
        sheet.delete_rows(1, sheet.max_row)
        header = ['Student ID', 'Name', 'Face Status']
        sheet.append(header)
        for row_id in table.get_children():
            row = table.item(row_id)['values']
            # print(row)
            sheet.append(row)            
        workbook.save(file_path)
        workbook.close()
        frame.destroy()
        std_page.StudentDataPage(window)
        
    # Create Button
    def createButton(frame, text, command):
        button = ctk.CTkButton(master=frame, fg_color='gray',
                                width=200, height=50,
                                text=text,
                                font=("Leelawadee", 25),
                                command=command)
        return button
    
    # Create Entry
    def createEntry(frame, wid, hold_text):
        entry = ctk.CTkEntry(frame, width=wid,
                             font=('Leelawadee', 20), text_color='black',
                             placeholder_text=hold_text,
                             fg_color="#f0f0f0")
        return entry
    
    # Select record Function
    def select_record(table, id_entry, name_entry):
        # Clear entry boxs
        id_entry.delete(0, 'end')
        name_entry.delete(0, 'end')
        # Grab record
        selected = table.focus()
        values = table.item(selected, 'values')
        # Output to boxs
        id_entry.insert(0, values[0])
        name_entry.insert(0, values[1])
    
    # Insert Data Function
    def insert_record(table, id_entry, name_entry):
        if id_entry.get() and name_entry.get() != '':
            table.insert(parent='', index='end', text="",
                        values=(id_entry.get(), name_entry.get(), 'None'))
            # Clear the boxes
            id_entry.delete(0, 'end')
            name_entry.delete(0, 'end')
        else:
            pass           
    
    # Update Record Function
    def update_record(table, id_entry, name_entry):
        if id_entry.get() and name_entry.get() != '':
            # Grab record number
            selected = table.focus()
            face_status = table.item(selected, 'values')[2]
            # Save new data
            table.item(selected, text="", values=(id_entry.get(), name_entry.get(), face_status))
        else:
            pass
    
    # Delete Record Function
    def delete_record(table):
        if table.selection() != () :   
            for item in table.selection():
                table.delete(item)
        else:
            pass
        
    # Create Frame 1 Function
    def createFrame1(window, master_frame, file_path):
        # Load Excel File    
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        list_of_data = list(sheet.values)

        # Create Sub master Frame 1
        sub_master_frame_1 = ctk.CTkFrame(master=master_frame)
        sub_master_frame_1.grid(row=0, column=0, padx=(10,0), pady=10 )
        
        # Create Student Information Label
        student_information_label = ctk.CTkLabel(master=sub_master_frame_1,
                                                text="Student Information",
                                                font=("Leelawadee", 35, "bold"),
                                                padx=20, pady=10)
        student_information_label.pack(pady=(15, 0))
        
        # Create Student Date Frame
        student_data_frame = ctk.CTkFrame(master=sub_master_frame_1)
        student_data_frame.pack()

        # Create Table Frame
        table_frame = ctk.CTkFrame(master=student_data_frame)
        table_frame.grid(row=0, column=0, padx=10, pady=5)
        
        # Create Scrollbar and Table
        tree_scrollbar = ttk.Scrollbar(table_frame, orient="vertical")
        tree_scrollbar.pack(side="right", fill="y")
        table = ttk.Treeview(table_frame, height=20, show="headings",
                                     yscrollcommand=tree_scrollbar.set)
        table.pack(fill="both", expand=True)
        tree_scrollbar.config(command=table.yview)
        
        # Configure Table
        style = ttk.Style()
        style.configure("Treeview.Heading", font=("Leelawadee", 12, "bold"))
        style.configure("Treeview", font=("Leelawadee", 15), rowheight=20)

        # Insert Table
        table['columns'] = list_of_data[0]
        table.column('#0', width=0, stretch='YES')
        table.column('Student ID', anchor='center', width=130)
        table.column('Name', width=260)
        table.column('Face Status', anchor='center', width=100)
        for header in list_of_data[0]:
            table.heading(str(header), text=str(header))
        for value in list_of_data[1:]:
            table.insert("", "end", values=value)
            
        # Create Student Entry Frame
        entry_frame = ctk.CTkFrame(master=sub_master_frame_1)
        entry_frame.pack()
        
        # Create Entry for show selection data
        style.configure("Entry", font=('Leelawadee', 20, 'bold'))
        id_label = ctk.CTkLabel(entry_frame, text='Student ID')
        id_label.grid(row=0, column=0)
        id_entry = ShowStudentDataPage.createEntry(entry_frame, 150, 'Student ID')
        id_entry.grid(row=1, column=0)
        name_label = ctk.CTkLabel(entry_frame, text='Full Name')
        name_label.grid(row=0, column=1)
        name_entry = ShowStudentDataPage.createEntry(entry_frame, 350, 'Full Name')
        name_entry.grid(row=1, column=1)
        
        # Create Button Frame
        button_frame = ctk.CTkFrame(master=sub_master_frame_1)
        button_frame.pack(pady=(10,0))
        
        # Create Insert Data Button
        insert_data_button = ShowStudentDataPage.createButton(button_frame, 'Insert Data',
                                                              lambda: ShowStudentDataPage.insert_record(table, id_entry, name_entry))
        insert_data_button.grid(row=0, column=0, padx=20, pady=10)
        
        # Create Update Data Button
        update_data_button = ShowStudentDataPage.createButton(button_frame, 'Update Data',
                                                              lambda: ShowStudentDataPage.update_record(table, id_entry, name_entry))
        update_data_button.grid(row=1, column=0, padx=20, pady=10)
        
        # Create Delete Data Button
        delete_data_button = ShowStudentDataPage.createButton(button_frame, 'Delete Data',
                                                              lambda: ShowStudentDataPage.delete_record(table))
        delete_data_button.grid(row=2, column=0, padx=20, pady=10)
        
        # bind
        table.bind('<Delete>', lambda e: ShowStudentDataPage.delete_record(table))
        table.bind('<ButtonRelease-1>', lambda e: ShowStudentDataPage.select_record(table, id_entry, name_entry))
        
        # Back and save Button
        back_n_save_button = ctk.CTkButton(master=sub_master_frame_1, fg_color='green',
                                    width=200, height=50,
                                    text="Back and Save",
                                    font=("Leelawadee", 25),
                                    command=lambda: ShowStudentDataPage.backAndSave(window, master_frame, workbook, table, file_path))
        back_n_save_button.pack(side="bottom", pady=(30,30))
        
        # Back without save Button
        back_button = ctk.CTkButton(master=sub_master_frame_1, fg_color='red',
                                    width=200, height=50,
                                    text="Back without Save",
                                    font=("Leelawadee", 25),
                                    command=lambda: ShowStudentDataPage.back(window, master_frame))
        back_button.pack(side="bottom", pady=(30,0))
        
    # Create Camera Lebel and Show
    def capture_video(camid, cam_frame):
        width = 480
        height = 480
        cam = cv2.VideoCapture(camid+cv2.CAP_DSHOW)
        cam.set(cv2.CAP_PROP_FRAME_WIDTH, width)
        cam.set(cv2.CAP_PROP_FRAME_HEIGHT, height)
        lmain = ctk.CTkLabel(cam_frame, text="")
        lmain.pack()
        def show_frame():
            _, frame = cam.read()
            cv2image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGBA)
            img = Image.fromarray(cv2image)
            imgctk = ctk.CTkImage(light_image=img, dark_image=img, size=img.size)
            lmain.configure(image=imgctk)
            lmain.after(20, show_frame)
            
        show_frame()
        
    def collect_face():
        print("Collecting Face")
    
    # Create Frame 2 Function
    def createFrame2(window, master_frame):
        # Create Sub Master Frame 2
        sub_master_frame_2 = ctk.CTkFrame(master=master_frame)
        sub_master_frame_2.grid(row=0, column=1, padx=5, pady=5)
        
        # Create Camera Main Frame
        camera_main_frame = ctk.CTkFrame(master=sub_master_frame_2)
        camera_main_frame.grid(row=0, column=0, padx=5, pady=5)
        
        # Create Camera 1 Frame
        camera_1_frame = ctk.CTkFrame(master=camera_main_frame)
        camera_1_frame.grid(row=0, column=0, padx=5, pady=5)
        
        # Create Camera 2 Frame
        camera_2_frame = ctk.CTkFrame(master=camera_main_frame)
        camera_2_frame.grid(row=0, column=1, padx=5, pady=5)
        
        # Start Thread
        thread1 = CameraThread(0, camera_1_frame)
        thread2 = CameraThread(1, camera_2_frame)
        thread1.start()
        thread2.start()
        
                # Create Facecollection Frame
        collect_button_frame = ctk.CTkFrame(master=sub_master_frame_2)
        collect_button_frame.grid(row=1, column=0, padx=5, pady=5)
        
        # Create Facecollection Button
        collect_button = ShowStudentDataPage.createButton(collect_button_frame, 'Capture Face',
                                                              lambda: ShowStudentDataPage.collect_face())
        collect_button.grid(row=0, column=0, padx=10, pady=10)
    
    def __init__(self, window, file_path):
        
        # Create Master Frame
        master_frame = ctk.CTkFrame(master=window)
        master_frame.pack(fill="both", expand=True)
        master_frame.grid_rowconfigure(0, weight=1)
        master_frame.grid_columnconfigure(0, weight=1)
        master_frame.grid_columnconfigure(1, weight=1)
        
        ShowStudentDataPage.createFrame1(window, master_frame, file_path)
        ShowStudentDataPage.createFrame2(window, master_frame)