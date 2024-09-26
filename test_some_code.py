import openpyxl


filename = "E:/Code-Thesis/version.0.5/students_data/2567.xlsx"

wb = openpyxl.load_workbook(filename)
sheetname = wb.sheetnames[0]
ws = wb[sheetname]

indx = []
for i in range(len(tuple(ws.rows))):
#             print(tuple(ws.rows)[i])
    flag = False
    for cell in tuple(ws.rows)[i]:
        if cell.value != None:
            flag = True
            break
    if flag == False:
        indx.append(i)
indx.sort()
#         print(filename)
for i in range(len(indx)):
#             print(indx[i]+1, "-",i,"=",indx[i]+1-i)
    
    ws.delete_rows(idx = indx[i]+1-i)

wb.save(filename)
wb.close()

# String dataDir = Utils.getDataDir(DeletingBlankRows.class);
# // Create a new Workbook. Open an existing excel file.
# Workbook wb = new Workbook(dataDir + "Book1.xlsx");

# // Create a Worksheets object with reference to the sheets of the Workbook.
# WorksheetCollection sheets = wb.getWorksheets();

# // Get first Worksheet from WorksheetCollection
# Worksheet sheet = sheets.get(0);

# // Delete the Blank Rows from the worksheet
# sheet.getCells().deleteBlankRows();

# // Save the excel file.
# wb.save(dataDir + "Output.xlsx");