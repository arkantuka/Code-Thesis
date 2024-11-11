import customtkinter as ctk
import datetime as dt
import openpyxl
import tkcalendar as cal
import os
import pages.manage_course_menu as mng_crs_menu
from CTkMessagebox import CTkMessagebox
from tktimepicker import AnalogPicker, AnalogThemes
from tkinter import filedialog

font = "THSarabunNew"
button_font_size = 30
header_font_size = 50
label_font_size = 20
entry_font_size = 20

class Clock:
    def updateTime(window, time, label, key):
        if key == "midterm":
            if time[2] == "AM":
                hour = time[0] if time[0] < 12 else time[0] - 12
                minute = time[1]
                AddCoursePage.midterm_start_time = "{}:{}".format(hour, minute)
            else:
                hour = time[0] + 12
                minute = time[1]
                AddCoursePage.midterm_start_time = "{}:{}".format(hour, minute)
        else:
            if time[2] == "AM":
                hour = time[0] if time[0] < 12 else time[0] - 12
                minute = time[1]
                AddCoursePage.final_start_time = "{}:{}".format(hour, minute)
            else:
                hour = time[0] + 12 if time[0] < 12 else time[0]
                minute = time[1]
                AddCoursePage.final_start_time = "{}:{}".format(hour, minute)
        label.configure(text="{}:{}".format(hour, minute))
        window.after(200, lambda: window.destroy()) 
        
    def __init__(self, midterm_start_time_label, key):
        root = ctk.CTk()
        root.title('Time Picker')
        time_picker = AnalogPicker(root)
        time_picker.pack(expand=True, fill="both")        
        theme = AnalogThemes(time_picker)
        theme.setDracula()
        submit = ctk.CTkButton(root, text="Submit", command=lambda: Clock.updateTime(root, time_picker.time(), midterm_start_time_label, key))
        submit.pack()
        
        root.mainloop()

class AddCoursePage:
    file_path = None
    midterm_start_time = None
    final_start_time = None
    # Create Back Function
    def back(window, master_frame):
        master_frame.destroy()
        mng_crs_menu.ManageCourseMenu(window)
        
    # Create Label
    def createLabel(frame, label_name):
        label = ctk.CTkLabel(master=frame,
                             text=label_name,
                             font=(font, label_font_size, "bold"),
                             corner_radius=8)
        return label
        
    # Create Entry
    def createEntry(frame, hold_text, width):
        entry = ctk.CTkEntry(master=frame, width=width,
                             font=(font, entry_font_size, "bold"), text_color='black',
                             placeholder_text=hold_text,
                             fg_color="#f0f0f0")
        return entry
    
    def addData(window, frame, courseID, courseName, term, year, midtermDate, finalDate, midtermTime, finalTime, file_path):
        midtermDate = midtermDate.strftime("%d/%m/%Y")
        finalDate = finalDate.strftime("%d/%m/%Y")
        data = {"Course ID" :courseID,"Course Name": courseName,
                "Term" : term,"Year" : year, 
                "Midterm Exam" : midtermDate, "Final Exam" : finalDate,
                "Midterm Start Time" : midtermTime, "Final Start Time" : finalTime}
        if None not in data.values():
            parent_path = "version.final/course_data"
            year_path = os.path.join(parent_path, year)
            if not os.path.exists(year_path): os.mkdir(year_path)
            path = os.path.join(year_path, term)
            if not os.path.exists(path): os.mkdir(path)
            
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.sheetnames
            new_workbook = openpyxl.Workbook()
            new_workbook.create_sheet(sheet[0])
            new_workbook.remove(new_workbook['Sheet'])
            new_workbook[sheet[0]].append(list(data.keys()))
            new_workbook[sheet[0]].append(list(data.values()))
            for row in workbook[sheet[0]].iter_rows(values_only=True):
                row = list(row)
                row[0] = str(row[0])
                row[0] = row[0].replace("\xa0", "")
                new_workbook[sheet[0]].append(row)
            new_workbook.save(f"{path}/{courseID}-{year}-{term}-{courseName}.xlsx")
            new_workbook.close()
            workbook.close()

            # Show Message
            CTkMessagebox(title="Information",
                        message=f"Course {courseID} added successfully.",
                        icon="check", option_1="OK", option_2=None)
            window.after(100, lambda: AddCoursePage.back(window, frame))
        else:
            CTkMessagebox(title="Error",
                        message="Please fill all the fields.",
                        icon="cancel", option_1="OK", option_2=None)     
    
    def fileDialog(initial_directory):
        file_path = filedialog.askopenfilename(initialdir=initial_directory,
                                        title="Select A File",
                                        filetype=(("xlsx files", "*.xlsx"),("All Files", "*.*")))
        return file_path
        
    def importData(label):
        try:
            file_path = AddCoursePage.fileDialog("/")
            label.configure(text=file_path)
            AddCoursePage.file_path = file_path
        except:
            pass
    
    # Create Frame 1
    def createFrame1(window, master_frame):
        # Create Sub master frame 1
        sub_master_frame_1 = ctk.CTkFrame(master=master_frame)
        sub_master_frame_1.grid(row=0, column=0, padx=(10,0), pady=10)
        sub_master_frame_1.grid_columnconfigure(0, weight=1)
        sub_master_frame_1.grid_columnconfigure(1, weight=1)
        
        main_label = ctk.CTkLabel(master=sub_master_frame_1,
                                   text="Add New Course Data",
                                   font=(font, header_font_size, "bold"),
                                   padx=20, pady=10)
        main_label.pack(pady=15)
        
        # Create Add Course Data Frame
        add_data_frame = ctk.CTkFrame(master=sub_master_frame_1)
        add_data_frame.pack(padx=10, pady=10)
        
        course_frame = ctk.CTkFrame(master=add_data_frame)
        course_frame.pack(padx=10, pady=10)
        
        # Create Course ID Label and Entry
        courseID_label = AddCoursePage.createLabel(course_frame, "Course ID")
        courseID_label.grid(row=0, column=0, padx=5, pady=0)
        courseID_entry = AddCoursePage.createEntry(course_frame, "Enter Course ID", 150)
        courseID_entry.grid(row=1, column=0, padx=5, pady=5)
        
        # Create Course Name Label and Entry
        courseName_label = AddCoursePage.createLabel(course_frame, "Course Name")
        courseName_label.grid(row=0, column=1, padx=5, pady=0)
        courseName_entry = AddCoursePage.createEntry(course_frame, "Enter Course Name", 400)
        courseName_entry.grid(row=1, column=1, padx=5, pady=5)
        
        # Create Radio Button Frame
        radio_frame = ctk.CTkFrame(master=add_data_frame)
        radio_frame.pack(padx=10, pady=10)
        radio_label = AddCoursePage.createLabel(radio_frame, "Select Term: ")
        radio_label.grid(row=0, column=0, padx=5, pady=0)
        radio_var = ctk.StringVar(master=radio_frame, value="1")
        radio_button_1 = ctk.CTkRadioButton(master=radio_frame, text="First Term",
                                            font=(font, entry_font_size, "bold"), 
                                            variable=radio_var, value="1")
        radio_button_1.grid(row=0, column=1, padx=5, pady=0)
        radio_button_2 = ctk.CTkRadioButton(master=radio_frame, text="Second Term",
                                            font=(font, entry_font_size, "bold"),
                                            variable=radio_var, value="2")
        radio_button_2.grid(row=0, column=2, padx=5, pady=0)
        radio_button_3 = ctk.CTkRadioButton(master=radio_frame, text="Summer Term",
                                            font=(font, entry_font_size, "bold"),
                                            variable=radio_var, value="3")
        radio_button_3.grid(row=0, column=3, padx=5, pady=0)
        
        # Create Combobox
        combobox_frame = ctk.CTkFrame(master=add_data_frame)
        combobox_frame.pack(padx=10, pady=10)
        combobox_label = AddCoursePage.createLabel(combobox_frame, "Select Year: ")
        combobox_label.grid(row=0, column=0, padx=5, pady=0)
        current_year = (dt.datetime.now().year)+543
        combobox_var = ctk.StringVar(master=combobox_frame, value=str(current_year))
        combobox = ctk.CTkComboBox(master=combobox_frame, 
                                values=[str(current_year-8),str(current_year-7), 
                                     str(current_year-6), str(current_year-5),  
                                     str(current_year-4), str(current_year-3), 
                                     str(current_year-2), str(current_year-1),
                                     str(current_year), str(current_year+1), 
                                     str(current_year+2)],
                                     font=(font, entry_font_size),variable=combobox_var)
        combobox.grid(row=0, column=1, padx=5, pady=0)
        
        # Create Date Time Picker
        date_time_picker_frame = ctk.CTkFrame(master=add_data_frame)
        date_time_picker_frame.pack(padx=10, pady=10)
        
        date_frame = ctk.CTkFrame(master=date_time_picker_frame)
        date_frame.grid(row=0, column=0, columnspan=2)
        midterm_date_frame = ctk.CTkFrame(master=date_frame)
        midterm_date_frame.grid(row=0, column=0, padx=(0,10))
        midterm_date_picker_label = AddCoursePage.createLabel(midterm_date_frame, "Midterm Exam Date: ")
        midterm_date_picker_label.grid(row=0, column=0)
        midterm_date_picker = cal.DateEntry(midterm_date_frame, date_pattern="dd-mm-yyyy", state="readonly", font=(font, 15))
        midterm_date_picker.grid(row=1, column=0)
        finale_date_frame = ctk.CTkFrame(master=date_frame)
        finale_date_frame.grid(row=0, column=1, padx=(10,0))
        final_date_picker_label = AddCoursePage.createLabel(finale_date_frame, "Final Exam Date: ")
        final_date_picker_label.grid(row=0, column=1)
        final_date_picker = cal.DateEntry(finale_date_frame, date_pattern="dd-mm-yyyy", state="readonly", font=(font, 15))
        final_date_picker.grid(row=1, column=1)
        
        # Create Time Clock
        time_frame = ctk.CTkFrame(master=date_time_picker_frame)
        time_frame.grid(row=1, column=0, padx=5, pady=(0,10))
        midterm_time_frame = ctk.CTkFrame(master=time_frame)
        midterm_time_frame.grid(row=0, column=0)
        midterm_start_time_label = ctk.CTkLabel(master=midterm_time_frame, text="Midterm Exam Start Time: ", font=(font, entry_font_size, "bold"))
        midterm_start_time_label.grid(row=0, column=0)
        midterm_show_time = ctk.CTkLabel(master=midterm_time_frame, text="00:00:00", font=(font, entry_font_size, "bold"))
        midterm_show_time.grid(row=1, column=0)
        midterm_start_btn = ctk.CTkButton(master=midterm_time_frame, text="Set Time", font=(font, entry_font_size, "bold"),
                            command=lambda: Clock(midterm_show_time, "midterm"))
        midterm_start_btn.grid(row=2, column=0)
        final_time_frame = ctk.CTkFrame(master=time_frame)
        final_time_frame.grid(row=0, column=1, padx=5, pady=(10,10))
        final_start_time_label = ctk.CTkLabel(master=final_time_frame, text="Final Exam Start Time: ", font=(font, entry_font_size, "bold"))
        final_start_time_label.grid(row=0, column=0)
        final_show_time = ctk.CTkLabel(master=final_time_frame, text="00:00:00", font=(font, entry_font_size, "bold"))
        final_show_time.grid(row=1, column=0)
        final_start_btn = ctk.CTkButton(master=final_time_frame, text="Set Time", font=(font, entry_font_size, "bold"),
                            command=lambda: Clock(final_show_time, "final"))
        final_start_btn.grid(row=2, column=0)
        
        # Create Choose File Frame
        choose_file_frame = ctk.CTkFrame(master=add_data_frame)
        choose_file_frame.pack(padx=10, pady=10)
        choose_file_label = AddCoursePage.createLabel(choose_file_frame, "Chosen File Path: ")
        choose_file_label.grid(row=0, column=0)
        path_show = ctk.CTkLabel(master=choose_file_frame, text="No File Chosen", font=(font, entry_font_size, "bold"))
        path_show.grid(row=1, column=0)

        # Create Buttons Frame
        button_frame = ctk.CTkFrame(master=sub_master_frame_1)
        button_frame.pack(pady=(15,0))
        
        # Create Import Button
        import_button = ctk.CTkButton(master=button_frame, fg_color='gray',
                                width=200, height=50,
                                text="Import Student Data",
                                font=(font, button_font_size, "bold"),
                                command=lambda: AddCoursePage.importData(path_show))
        import_button.grid(row=0, column=0, padx=10, pady=10)
        
        # Create Add Button
        submit_button_frame = ctk.CTkFrame(master=sub_master_frame_1)
        submit_button_frame.pack(pady=(20,0))
        submit_button = ctk.CTkButton(master=submit_button_frame, fg_color='green',
                                width=200, height=50,
                                text="Submit",
                                font=(font, button_font_size, "bold"),
                                command=lambda: AddCoursePage.addData(window, master_frame, 
                                                                      courseID_entry.get(), courseName_entry.get(), 
                                                                      radio_var.get(), combobox_var.get(), 
                                                                      midterm_date_picker.get_date(), final_date_picker.get_date(),
                                                                      AddCoursePage.midterm_start_time, AddCoursePage.final_start_time,
                                                                      AddCoursePage.file_path))
        submit_button.grid(row=0, column=0, padx=15, pady=15)
        
        # Exit Button
        exit_button = ctk.CTkButton(master=sub_master_frame_1, fg_color='red',
                                width=200, height=50,
                                text="Exit",
                                font=(font, button_font_size, "bold"),
                                command=window.destroy)
        exit_button.pack(side="bottom", pady=20)
        
        # Back Button
        back_button = ctk.CTkButton(master=sub_master_frame_1, fg_color='gray',
                                width=200, height=50,
                                text="Back",
                                font=(font, button_font_size, "bold"),
                                command=lambda: AddCoursePage.back(window, master_frame))
        back_button.pack(side="bottom", pady=(50,0))
    
    # Create Frame 2
    def createFrame2(window, master_frame):
        # Create Sub master frame 2
        sub_master_frame_2 = ctk.CTkFrame(master=master_frame)
        sub_master_frame_2.grid(row=0, column=1, padx=(0,10), pady=10 )
        
    def __init__(self, window):
        # Create Master Frame
        master_frame = ctk.CTkFrame(master=window)
        master_frame.pack(fill="both", expand=True)
        master_frame.grid_rowconfigure(0, weight=1)
        master_frame.grid_columnconfigure(0, weight=1)

        AddCoursePage.createFrame1(window, master_frame)
