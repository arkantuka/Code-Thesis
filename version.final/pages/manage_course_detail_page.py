import os
import customtkinter as ctk
import openpyxl
import openpyxl.workbook
import pages.manage_course_menu as mng_crs_menu
from tkinter import ttk
import tkcalendar as cal
from CTkSpinbox import *
from CTkMessagebox import CTkMessagebox

font = "THSarabunNew"
button_font_size = 30
header_font_size = 50
entry_font_size = 20
table_font_size = 15

class CourseDetail:
    # Back Button Function
    def back(window, frame):
        frame.destroy()
        mng_crs_menu.ManageCourseMenu(window)
        
    # Back and save Button Function
    def backAndSave(window, frame, workbook, table, file_path, dict_of_data):
        parent_dir = 'version.final/course_data'
        old_file_name = str(file_path).split('/')[-1]
        year = dict_of_data['Year']
        term = dict_of_data['Term']
        file_name = f"{dict_of_data['Course ID']}-{year}-{term}-{dict_of_data['Course Name']}.xlsx"
        new_path = os.path.join(parent_dir,year,term,file_name)
        sheetname = workbook.sheetnames
        sheet = workbook[sheetname[0]]
        sheet.delete_rows(1, sheet.max_row)
        sheet.append(list(dict_of_data.keys()))
        sheet.append(list(dict_of_data.values()))
        header = ['Student ID', 'Name', 'Face Status']
        sheet.append(header)
        for row_id in table.get_children():
            row = table.item(row_id)['values'][:3]
            sheet.append(row)            
        workbook.save(new_path)
        workbook.close()
        frame.destroy()
        mng_crs_menu.ManageCourseMenu(window)
        if old_file_name != file_name:
            os.remove(file_path)        
    
    def find_index_of_ids_duplicate(ids_in_database, ids_in_course):
        idx = []
        for id in ids_in_course:
            if id in ids_in_database:
                idx.append(ids_in_course.index(id)+3)
        return idx
    
    def update_face_status(ids):
        collection_ids = []
        year_path = os.listdir("version.final/images")
        for year in year_path:
            faceID_path = os.listdir(f"version.final/images/{year}")
            for id in faceID_path:
                collection_ids.append(id)
        return CourseDetail.find_index_of_ids_duplicate(collection_ids, ids)
    
    # Create Button
    def createButton(frame, text, command):
        button = ctk.CTkButton(master=frame, fg_color='gray',
                                width=200, height=50,
                                text=text,
                                font=(font, button_font_size, "bold"),
                                command=command)
        return button
    
    # Create Entry
    def createEntry(frame, wid, hold_text):
        entry = ctk.CTkEntry(frame, width=wid,
                             font=(font, entry_font_size), text_color='black',
                             placeholder_text=hold_text,
                             fg_color="#f0f0f0")
        return entry
    
    # Select record Function
    def select_record(table, id_entry, name_entry):
        try:
            # Clear entry boxs
            id_entry.delete(0, 'end')
            name_entry.delete(0, 'end')
            # Grab record
            selected = table.focus()
            values = table.item(selected, 'values')
            # Output to boxs
            id_entry.insert(0, values[0])
            name_entry.insert(0, values[1])
        except:
            pass        
        
    # Insert Data Function
    def add(table, id_entry, name_entry, all_ids, all_names):
        if id_entry.get() and name_entry.get() != '':
            if id_entry.get() in all_ids:
                CTkMessagebox(title='Error', message='ID already exists')
            elif name_entry.get() in all_names:
                CTkMessagebox(title='Error', message='Name already exists')
            else:
                table.insert(parent='', index='end', text="",
                            values=(id_entry.get(), name_entry.get(), 'None'))
                all_ids.append(id_entry.get())
                all_names.append(name_entry.get())
                # Clear the boxes
                id_entry.delete(0, 'end')
                name_entry.delete(0, 'end')
        else:
            pass
        
    # Update Record Function
    def edit(table, id_entry, name_entry, all_ids, all_names):
        if id_entry.get() and name_entry.get() != '':
            # Grab record number
            selected = table.focus()
            # Remove old data and Add new
            old_id = table.item(selected, 'values')[0]
            old_name = table.item(selected, 'values')[1]
            all_ids.remove(old_id)
            all_names.remove(old_name)
            all_ids.append(id_entry.get())
            all_names.append(name_entry.get())
            
            face_status = table.item(selected, 'values')[2]
            # Save new data
            table.item(selected, text="", values=(id_entry.get(), name_entry.get(), face_status))

            # Clear entry boxes
            id_entry.delete(0, 'end')
            name_entry.delete(0, 'end')
        else:
            pass
        
    # Delete Record Function
    def delete(table, id_entry, name_entry, all_ids, all_names):
        if table.selection() != () :
            selected = table.focus()
            values = table.item(selected, 'values')
            all_ids.remove(values[0])    
            all_names.remove(values[1])   
            for item in table.selection():
                table.delete(item)
                
            # Clear entry boxes
            id_entry.delete(0, 'end')
            name_entry.delete(0, 'end')
        else:
            pass
        
    def sort_column(table, col, reverse):
        data = [(table.set(item, col),item) for item in table.get_children('')]
        data.sort(reverse=reverse)
        for index, (val, item) in enumerate(data):
            table.move(item, '', index)
        table.heading(col, command=lambda: CourseDetail.sort_column(table, col, not reverse))
        
    def create_frame(window, master_frame, file_path):
        # Load Excel Data
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        loadToList = list(sheet.values)
        list_of_data = []
        for row in loadToList:
            list_of_data.append(list(row))
        dict_of_data = {list_of_data[0][i]: list_of_data[1][i] for i in range(len(list_of_data[0]))}
        list_of_data[2][2] = 'Face Status'
        allIDs = [str(i[0]) for i in list_of_data[3:]]
        allNames = [str(i[1]) for i in list_of_data[3:]]
        duplicate_index = CourseDetail.update_face_status(allIDs)
        for i in duplicate_index:
            list_of_data[i][2] = 'Collected'
        
        # Main Label
        main_label = ctk.CTkLabel(master=master_frame,
                                  text="Course Details",
                                  font=(font, header_font_size, "bold"),
                                  padx=100, pady=10)
        main_label.pack(pady=(15, 0))
        
        # Create Main Frame
        main_frame = ctk.CTkFrame(master=master_frame)
        main_frame.grid_columnconfigure(0, weight=1)
        main_frame.grid_columnconfigure(1, weight=1)
        main_frame.grid_columnconfigure(2, weight=1)
        main_frame.pack()
        
        # Create Frame Right
        frame_center = ctk.CTkFrame(master=main_frame)
        frame_center.grid(row=0, column=1, padx=(10, 0), pady=10)
        
        # Create Label
        student_label = ctk.CTkLabel(master=frame_center,
                                     text="Student Information",
                                     font=(font, button_font_size, "bold"),
                                     padx=10, pady=10)
        student_label.pack()
        
        # Create course detail frame
        course_detail_frame = ctk.CTkFrame(master=frame_center)
        course_detail_frame.pack(pady=10)
        
        # Create table frame
        table_frame = ctk.CTkFrame(master=course_detail_frame)
        table_frame.pack()
        
        # Create Scrollbar and Table
        tree_scrollbar = ttk.Scrollbar(table_frame, orient="vertical")
        tree_scrollbar.pack(side="right", fill="y")
        table = ttk.Treeview(table_frame, height=20, show="headings",
                                     yscrollcommand=tree_scrollbar.set)
        table.pack(fill="both", expand=True)
        tree_scrollbar.config(command=table.yview)
        
        # Configure Table
        style = ttk.Style()
        style.configure("Treeview.Heading", font=(font, 15, "bold"))
        style.configure("Treeview", font=(font, 15), rowheight=27)

        # Insert Table
        table['columns'] = list_of_data[2][:3]
        table.column('#0', width=0, stretch='YES')
        table.column('Student ID', anchor='center', width=130)
        table.column('Name', width=260)
        table.column('Face Status', anchor='center', width=100)
        for header in list_of_data[2][:3]:
            table.heading(header, text=header, command=lambda c=header: CourseDetail.sort_column(table, c, False))
        for value in list_of_data[3:]:
            table.insert("", "end", values=value)
            
        # Create Student Entry Frame
        entry_frame = ctk.CTkFrame(master=frame_center)
        entry_frame.pack()
        
        # Create Entry for show selection data
        style.configure("Entry", font=(font, entry_font_size, 'bold'))
        id_label = ctk.CTkLabel(entry_frame, text='Student ID')
        id_label.grid(row=0, column=0)
        id_entry = CourseDetail.createEntry(entry_frame, 150, 'Student ID')
        id_entry.grid(row=1, column=0)
        name_label = ctk.CTkLabel(entry_frame, text='Full Name')
        name_label.grid(row=0, column=1)
        name_entry = CourseDetail.createEntry(entry_frame, 350, 'Full Name')
        name_entry.grid(row=1, column=1)
        
        # Create Right Frame
        frame_right = ctk.CTkFrame(master=main_frame)
        frame_right.grid(row=0, column=2, padx=(10, 0), pady=10)
        
        # Create Course Description Label
        course_description_label = ctk.CTkLabel(master=frame_right,
                                                text="Course Description",
                                                font=(font, 30, "bold"),
                                                padx=100, pady=10)
        course_description_label.pack()
        
        # Create Course Description Frame
        course_description_frame = ctk.CTkFrame(master=frame_right)
        course_description_frame.pack()
        
        # Create Information Frame
        courseID_label = ctk.CTkLabel(master=course_description_frame,
                                     text="Course ID: "+dict_of_data['Course ID'],
                                     font=(font, 20))
        courseID_label.pack()
        courseID_var = ctk.StringVar()
        courseID_var.set(dict_of_data['Course ID'])
        courseID_entry = ctk.CTkEntry(master=course_description_frame,
                                      fg_color="#f0f0f0",
                                      text_color='black',
                                      textvariable=courseID_var,
                                      width=220, font=(font, 20))
        courseID_entry.pack(pady=(0, 10))
        
        courseName_label = ctk.CTkLabel(master=course_description_frame,
                                       text="Course Name: "+dict_of_data['Course Name'],
                                       font=(font, 20))
        courseName_label.pack()
        courseName_var = ctk.StringVar()
        courseName_var.set(dict_of_data['Course Name'])
        courseName_entry = ctk.CTkEntry(master=course_description_frame,
                                        fg_color="#f0f0f0",
                                        text_color='black',
                                        textvariable=courseName_var,
                                        width=220, font=(font, 20))
        courseName_entry.pack(pady=(0, 10))
        
        courseTerm_label = ctk.CTkLabel(master=course_description_frame,
                                       text="Term: "+dict_of_data['Term'],
                                       font=(font, 20))
        courseTerm_label.pack()
        radio_frame = ctk.CTkFrame(master=course_description_frame)
        radio_frame.pack(pady=(0, 10))
        radio_var = ctk.StringVar()
        radio_var.set(dict_of_data['Term'])
        radio_button1 = ctk.CTkRadioButton(master=radio_frame, variable=radio_var, value="1", text="First Term")
        radio_button1.grid(row=0, column=0, padx=5, pady=0)
        radio_button2 = ctk.CTkRadioButton(master=radio_frame, variable=radio_var, value="2", text="Second Term")
        radio_button2.grid(row=0, column=1, padx=5, pady=0)
        radio_button3 = ctk.CTkRadioButton(master=radio_frame, variable=radio_var, value="3", text="Summer Term")
        radio_button3.grid(row=0, column=2, padx=5, pady=0)
        
        courseYear_label = ctk.CTkLabel(master=course_description_frame,
                                       text="Year: "+dict_of_data['Year'],
                                       font=(font, 20))
        courseYear_label.pack()
        courseYear_var = ctk.IntVar()
        courseYear_var.set(dict_of_data['Year'])
        courseYear_entry = ctk.CTkEntry(master=course_description_frame,
                                        fg_color="#f0f0f0",
                                        text_color='black',
                                        textvariable=courseYear_var,
                                        width=220, font=(font, 20))
        courseYear_entry.pack(pady=(0, 10))
        
        course_midterm_date_label = ctk.CTkLabel(master=course_description_frame,
                                                text="Midterm Date: "+dict_of_data['Midterm Exam'],
                                                font=(font, 20))
        course_midterm_date_label.pack()
        midterm_date_frame = ctk.CTkFrame(master=course_description_frame)
        midterm_date_frame.pack(pady=(0, 10))
        midterm_date_picker = cal.DateEntry(master=midterm_date_frame, date_pattern='dd-mm-yyyy',
                                            state='readonly', font=(font, 15))
        midterm_date_picker.set_date(dict_of_data['Midterm Exam'])
        midterm_date_picker.pack()
        
        course_midterm_time_label = ctk.CTkLabel(master=course_description_frame,
                                                text="Midterm Time: "+dict_of_data['Midterm Start Time'],
                                                font=(font, 20))
        course_midterm_time_label.pack()
        spinbox_midterm_frame = ctk.CTkFrame(master=course_description_frame)
        spinbox_midterm_frame.pack(pady=(0, 10))
        spinmid_hour_var = ctk.IntVar()
        spinmid_hour_var.set(dict_of_data['Midterm Start Time'].split(':')[0])
        spinbox_midterm_hour = CTkSpinbox(master=spinbox_midterm_frame, min_value=0, max_value=23, start_value=spinmid_hour_var.get(), variable=spinmid_hour_var)
        spinbox_midterm_hour.grid(row=0, column=0, padx=5, pady=0)
        spinmid_min_var = ctk.IntVar()
        spinmid_min_var.set(dict_of_data['Midterm Start Time'].split(':')[1])
        spinbox_midterm_minute = CTkSpinbox(master=spinbox_midterm_frame, min_value=0, max_value=59, start_value=spinmid_min_var.get(), variable=spinmid_min_var)
        spinbox_midterm_minute.grid(row=0, column=1, padx=5, pady=0)
        
        course_final_date_label = ctk.CTkLabel(master=course_description_frame,
                                               text="Final Date: "+dict_of_data['Final Exam'],
                                               font=(font, 20))
        course_final_date_label.pack()
        final_date_frame = ctk.CTkFrame(master=course_description_frame)
        final_date_frame.pack(pady=(0, 10))
        final_date_picker = cal.DateEntry(master=final_date_frame, date_pattern='dd-mm-yyyy',
                                          state='readonly', font=(font, 15))
        final_date_picker.set_date(dict_of_data['Final Exam'])
        final_date_picker.pack()
        
        course_final_time_label = ctk.CTkLabel(master=course_description_frame,
                                               text="Final Time: "+dict_of_data['Final Start Time'],
                                               font=(font, 20))
        course_final_time_label.pack()
        spinbox_final_frame = ctk.CTkFrame(master=course_description_frame)
        spinbox_final_frame.pack(pady=(0, 10))
        spinfinal_hour_var = ctk.IntVar()
        spinfinal_hour_var.set(dict_of_data['Final Start Time'].split(':')[0])
        spinbox_final_hour = CTkSpinbox(master=spinbox_final_frame, min_value=0, max_value=23, start_value=spinfinal_hour_var.get(), variable=spinfinal_hour_var)
        spinbox_final_hour.grid(row=0, column=0, padx=5, pady=0)
        spinfinal_min_var = ctk.IntVar()
        spinfinal_min_var.set(dict_of_data['Final Start Time'].split(':')[1])
        spinbox_final_minute = CTkSpinbox(master=spinbox_final_frame, min_value=0, max_value=59, start_value=spinfinal_min_var.get(), variable=spinfinal_min_var)
        spinbox_final_minute.grid(row=0, column=1, padx=5, pady=0)
        dict_of_data['Final Start Time'] = str(spinfinal_hour_var.get())+':'+str(spinfinal_min_var.get())

        def edit():
            dict_of_data['Course ID'] = courseID_entry.get()
            dict_of_data['Course Name'] = courseName_entry.get()
            dict_of_data['Term'] = radio_var.get()
            dict_of_data['Year'] = courseYear_entry.get()
            dict_of_data['Midterm Exam'] = midterm_date_picker.get_date().strftime("%d/%m/%Y")
            dict_of_data['Midterm Start Time'] = str(spinmid_hour_var.get())+':'+str(spinmid_min_var.get())
            dict_of_data['Final Exam'] = final_date_picker.get_date().strftime("%d/%m/%Y")
            dict_of_data['Final Start Time'] = str(spinfinal_hour_var.get())+':'+str(spinfinal_min_var.get())
            
        # Create Edit Buttons Frame
        edit_button = CourseDetail.createButton(course_description_frame, 'Edit', lambda: edit())
        edit_button.pack(pady=(20, 10))
        
        # Create Frame Left
        frame_left = ctk.CTkFrame(master=main_frame)
        frame_left.grid(row=0, column=0, padx=(0, 10), pady=10)
        
        # Create Buttons Frame
        buttons_frame = ctk.CTkFrame(master=frame_left)
        buttons_frame.pack(pady=10)
        
        # Create Add Button
        add_button = CourseDetail.createButton(buttons_frame, 'Add', lambda: CourseDetail.add(table, id_entry, name_entry, allIDs, allNames))
        add_button.grid(row=0, column=0, padx=20, pady=10)
        
        # Create Edit Button
        edit_button = CourseDetail.createButton(buttons_frame, 'Edit', lambda: CourseDetail.edit(table, id_entry, name_entry, allIDs, allNames))
        edit_button.grid(row=1, column=0, padx=20, pady=10)
        
        # Create Delete Button
        delete_button = CourseDetail.createButton(buttons_frame, 'Delete', lambda: CourseDetail.delete(table, id_entry, name_entry, allIDs, allNames))
        delete_button.grid(row=2, column=0, padx=20, pady=10)
        
        # bind
        table.bind('<Delete>', lambda e: CourseDetail.delete(table))
        table.bind('<ButtonRelease-1>', lambda e: CourseDetail.select_record(table, id_entry, name_entry))
        
        # Back and save Button
        back_n_save_button = ctk.CTkButton(master=frame_left, fg_color='green',
                                    width=220, height=50,
                                    text="Back and Save",
                                    font=(font, 25),
                                    command=lambda: CourseDetail.backAndSave(window, master_frame, workbook, table, file_path, dict_of_data))
        back_n_save_button.pack(side="bottom", pady=(30,30))
        
        # Back Button
        back_button = ctk.CTkButton(master=frame_left, fg_color='red',
                                    width=220, height=50,
                                    text="Back without save",
                                    font=(font, 25),
                                    command=lambda: CourseDetail.back(window, master_frame))
        back_button.pack(side="bottom", pady=(100, 0))
    
    def __init__(self, window, file_path):
        # Create Master Frame
        master_frame = ctk.CTkFrame(master=window)
        master_frame.pack(fill="both", expand=True)
        master_frame.grid_rowconfigure(0, weight=1)
        master_frame.grid_columnconfigure(0, weight=1)
        
        CourseDetail.create_frame(window, master_frame, file_path)