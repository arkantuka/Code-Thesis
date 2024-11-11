import threading
import customtkinter as ctk
import openpyxl
import cv2
import os
import pages.manage_student_data_menu_page as mng_std_menu
from PIL import Image
from tkinter import messagebox, ttk
from CTkMessagebox import CTkMessagebox

font = "THSarabunNew"
button_font_size = 30
header_font_size = 50
entry_font_size = 20
table_font_size = 15

class ManageStudentPage:
    
    student_id = ''
    student_name = ''
    faces_path = ''
    capture_status = False
    cam1_count = 0
    cam2_count = 0
    table = None
    
    # Back Button Function
    def back(window, frame, workbook):
        workbook.close()
        frame.destroy()
        mng_std_menu.StudentMenu(window)
    
    # Back and save Button Function
    def backAndSave(window, frame, workbook, table, file_path):
        sheetname = workbook.sheetnames
        sheet = workbook[sheetname[0]]
        sheet.delete_rows(1, sheet.max_row)
        header = ['Student ID', 'Name', 'Face Status']
        sheet.append(header)
        for row_id in table.get_children():
            row = table.item(row_id)['values']
            row[0] = str(row[0])
            sheet.append(row)            
        workbook.save(file_path)
        workbook.close()
        frame.destroy()
        mng_std_menu.StudentMenu(window)
        
    # Create Button
    def createButton(frame, text, command):
        button = ctk.CTkButton(master=frame, fg_color='gray',
                                width=200, height=50,
                                text=text,
                                font=(font, button_font_size, "bold"),
                                command=command)
        return button
    
    # Create Entry
    def createEntry(frame, wid, hold_text):
        entry = ctk.CTkEntry(frame, width=wid,
                             font=(font, entry_font_size), text_color='black',
                             placeholder_text=hold_text,
                             fg_color="#f0f0f0")
        return entry
    
    # Select record Function
    def select_record(table, id_entry, name_entry):
        try:
            # Clear entry boxs
            id_entry.delete(0, 'end')
            name_entry.delete(0, 'end')
            # Grab record
            selected = table.focus()
            values = table.item(selected, 'values')
            # Output to boxs
            id_entry.insert(0, values[0])
            name_entry.insert(0, values[1])
            # Set ID and Name
            ManageStudentPage.set_id_name(id_entry, name_entry)
        except:
            pass
        
    # Insert Data Function
    def insert_record(table, id_entry, name_entry):
        if id_entry.get() and name_entry.get() != '':
            table.insert(parent='', index='end', text="",
                        values=(id_entry.get(), name_entry.get(), 'None'))
            # Clear the boxes
            id_entry.delete(0, 'end')
            name_entry.delete(0, 'end')
        else:
            pass           
    
    def set_table(table):
        ManageStudentPage.table = table
        
    # Update Record Function
    def update_record(table, id_entry, name_entry):
        if id_entry.get() and name_entry.get() != '':
            # Grab record number
            selected = table.focus()
            face_status = table.item(selected, 'values')[2]
            # Save new data
            table.item(selected, text="", values=(id_entry.get(), name_entry.get(), face_status))
        else:
            pass
    
    # Delete Record Function
    def delete_record(table):
        if table.selection() != () :   
            for item in table.selection():
                table.delete(item)
        else:
            pass
        
    # Set ID and Name
    def set_id_name(id_entry, name_entry):
        if id_entry.get() and name_entry.get() != '':
            ManageStudentPage.student_id = id_entry.get()
            ManageStudentPage.student_name = name_entry.get()
        else:
            pass
    
    def find_ids_index_duplicate(ids_database, ids_opened):
        idx = []
        for id in ids_database:
            if id in ids_opened:
                idx.append(ids_opened.index(id)+3)
        return idx
    
    def update_face_status(ids, year_path):
        collection_ids = []
        faceID_path = os.listdir(f"version.final/images/{year_path}")
        for id in faceID_path:
            collection_ids.append(id)
        return ManageStudentPage.find_ids_index_duplicate(collection_ids, ids)
    
    def sort_column(table, col, reverse):
        data = [(table.set(item, col),item) for item in table.get_children('')]
        data.sort(reverse=reverse)
        for index, (val, item) in enumerate(data):
            table.move(item, '', index)
        table.heading(col, command=lambda: ManageStudentPage.sort_column(table, col, not reverse))
            
    # Create Frame 1 Function
    def createFrame1(window, master_frame, file_path):
        # Load Excel File    
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        loadToList = list(sheet.values)
        list_of_data = []
        for row in loadToList:
            if row[0][:] == 'None':
                continue
            list_of_data.append(list(row))
        all_ids = [i[0] for i in list_of_data[3:]]
        year_path = os.path.basename(file_path)
        year_path = os.path.splitext(year_path)[0]
        duplicate_ids = ManageStudentPage.update_face_status(all_ids, year_path)
        for idx in duplicate_ids:
            list_of_data[idx][2] = "Collected"        

        # Create Sub master Frame 1
        sub_master_frame_1 = ctk.CTkFrame(master=master_frame)
        sub_master_frame_1.pack(padx=10, pady=10)
        
        # Create Student Information Label
        student_information_label = ctk.CTkLabel(master=sub_master_frame_1,
                                                text="Student Information",
                                                font=(font, header_font_size, "bold"),
                                                padx=20, pady=10)
        student_information_label.pack(pady=(15, 0))
        
        # Create Sub Frame
        sub_frame_1 = ctk.CTkFrame(master=sub_master_frame_1)
        sub_frame_1.pack(padx=10, pady=10)
        
        # Create Student Date Frame
        student_data_frame = ctk.CTkFrame(master=sub_frame_1)
        student_data_frame.grid(row=0, column=0, padx=10, pady=5)

        # Create Table Frame
        table_frame = ctk.CTkFrame(master=student_data_frame)
        table_frame.pack(fill="both", expand=True)
        
        # Create Scrollbar and Table
        tree_scrollbar = ttk.Scrollbar(table_frame, orient="vertical")
        tree_scrollbar.pack(side="right", fill="y")
        table = ttk.Treeview(table_frame, height=30, show="headings",
                                     yscrollcommand=tree_scrollbar.set)
        table.pack(fill="both", expand=True)
        tree_scrollbar.config(command=table.yview)
        
        # Configure Table
        style = ttk.Style()
        style.configure("Treeview.Heading", font=(font, table_font_size,"bold"))
        style.configure("Treeview", font=(font, table_font_size), rowheight=20)

        # Insert Table
        table['columns'] = list_of_data[0]
        table.column('#0', width=0, stretch='YES')
        table.column('Student ID', anchor='center', width=150)
        table.column('Name', width=300)
        table.column('Face Status', anchor='center', width=120)
        for header in list_of_data[0]:
            table.heading(str(header), text=str(header), command=lambda c=header: ManageStudentPage.sort_column(table, c, False))
        for value in list_of_data[1:]:
            table.insert("", "end", values=value)
            
        # Create Student Entry Frame
        entry_frame = ctk.CTkFrame(master=student_data_frame)
        entry_frame.pack()
        
        # Create Entry for show selection data
        style.configure("Entry", font=(font, entry_font_size, 'bold'))
        id_label = ctk.CTkLabel(entry_frame, text='Student ID')
        id_label.grid(row=0, column=0)
        id_entry = ManageStudentPage.createEntry(entry_frame, 150, 'Student ID')
        id_entry.grid(row=1, column=0)
        name_label = ctk.CTkLabel(entry_frame, text='Full Name')
        name_label.grid(row=0, column=1)
        name_entry = ManageStudentPage.createEntry(entry_frame, 350, 'Full Name')
        name_entry.grid(row=1, column=1)
        
        ManageStudentPage.set_table(table)
        
        # Create Button Frame
        button_frame = ctk.CTkFrame(master=sub_frame_1)
        button_frame.grid(row=0, column=1, padx=10, pady=5)
        
        # Create Insert Data Button
        insert_data_button = ManageStudentPage.createButton(button_frame, 'Insert Data',
                                                              lambda: ManageStudentPage.insert_record(table, id_entry, name_entry))
        insert_data_button.grid(row=0, column=0, padx=20, pady=10)
        
        # Create Update Data Button
        update_data_button = ManageStudentPage.createButton(button_frame, 'Update Data',
                                                              lambda: ManageStudentPage.update_record(table, id_entry, name_entry))
        update_data_button.grid(row=1, column=0, padx=20, pady=10)
        
        # Create Delete Data Button
        delete_data_button = ManageStudentPage.createButton(button_frame, 'Delete Data',
                                                              lambda: ManageStudentPage.delete_record(table))
        delete_data_button.grid(row=2, column=0, padx=20, pady=10)
        
        # bind
        table.bind('<Delete>', lambda e: ManageStudentPage.delete_record(table))
        table.bind('<ButtonRelease-1>', lambda e: ManageStudentPage.select_record(table, id_entry, name_entry))
        
        # Back and save Button
        back_n_save_button = ctk.CTkButton(master=button_frame, fg_color='green',
                                    width=220, height=50,
                                    text="Back and Save",
                                    font=(font, button_font_size, "bold"),
                                    command=lambda: ManageStudentPage.backAndSave(window, master_frame, workbook, table, file_path))
        back_n_save_button.grid(row=3, column=0, pady=(40,0))
        
        # Back without save Button
        back_button = ctk.CTkButton(master=button_frame, fg_color='red',
                                    width=220, height=50,
                                    text="Back without Save",
                                    font=(font, button_font_size, "bold"),
                                    command=lambda: ManageStudentPage.back(window, master_frame, workbook))
        back_button.grid(row=4, column=0, pady=(20,10))
    
    def __init__(self, window, file_path):
        
        # Create Master Frame
        master_frame = ctk.CTkFrame(master=window)
        master_frame.pack(fill="both", expand=True)
        
        year_path = os.path.basename(file_path)
        year_path = os.path.splitext(year_path)[0]
        
        ManageStudentPage.createFrame1(window, master_frame, file_path)