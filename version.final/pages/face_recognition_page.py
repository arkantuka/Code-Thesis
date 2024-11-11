import os
import customtkinter as ctk
import cv2
import datetime
import openpyxl
import pages.main_page as mp
from PIL import Image
from tkinter import ttk

font = "THSarabunNew"
button_font_size = 30
header_font_size = 50

class FaceRecognition:
    
    attendance_data = []
    
    # Back Button Function
    def back(window, frame):
        frame.destroy()
        cam1 = cv2.VideoCapture(0+cv2.CAP_DSHOW)
        cam1.release()
        mp.MainPage(window)
        
    def writeAttendance(old_path, index):
        course_name = str(old_path)[33:-5]
        parent_dir = 'version.final/time_attendance'
        file_path = os.path.join(parent_dir, course_name+'.xlsx')
        if not os.path.exists(file_path):
            wb = openpyxl.load_workbook(old_path)
            sheet = wb.active
            loadToList = list(sheet.values)
            list_of_data = []
            sheetname = wb.sheetnames
            sheet = wb[sheetname[0]]
            for row in loadToList:
                list_of_data.append(list(row))
            dict_of_data = {list_of_data[0][i]: list_of_data[1][i] for i in range(len(list_of_data[0]))}
            sheet.delete_rows(1, sheet.max_row)
            sheet.append(list(dict_of_data.keys()))
            sheet.append(list(dict_of_data.values()))
            header = ['Student ID', 'Name', 'Date', 'Time']
            sheet.append(header)
            for row_id in list_of_data[3:]:
                row = row_id[:2]
                sheet.append(row)
            wb.save("version.final/time_attendance"+"/"+str(old_path)[33:-5]+".xlsx")
        workbook = openpyxl.load_workbook(file_path)
        sheetname = workbook.sheetnames
        sheet = workbook[sheetname[0]]
        sheet.cell(row=index+4, column=3).value = datetime.datetime.now().strftime("%d/%m/%Y")
        sheet.cell(row=index+4, column=4).value = datetime.datetime.now().strftime("%H:%M:%S")
        workbook.save(file_path)
        workbook.close()
        
    def create_camera_frame(frame, file_path, studentIDs, studentNames):
        course_name = str(file_path)[33:-5]
        # Create Camera Frame
        cam_frame = ctk.CTkFrame(master=frame)
        cam_frame.grid(row=0, column=0, padx=20, pady=10)
        
        already_detected = []
        
        directory = os.path.dirname(__file__)
        weights = os.path.join(directory, "face_detection_yunet_2023mar.onnx")
        face_cascade_yunet = cv2.FaceDetectorYN_create(weights, "", (0, 0))
        
        width = 1280
        height = 890 
        cam = cv2.VideoCapture(0+cv2.CAP_DSHOW)
        cam.set(cv2.CAP_PROP_FRAME_WIDTH, width)
        cam.set(cv2.CAP_PROP_FRAME_HEIGHT, height)
        lmain = ctk.CTkLabel(cam_frame, text="")
        lmain.pack()
        face_recognizer = cv2.face.LBPHFaceRecognizer_create()
        model_path = 'version.final/training_models/'+course_name+'.yml'
        face_recognizer.read(model_path)
        
        def show_frame():
            _, frame = cam.read()
            cv2image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGBA)
            gray_img = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            
            h, w, _ = frame.shape
            faces_yn = face_cascade_yunet.setInputSize((w, h))
            _, faces_yn = face_cascade_yunet.detect(frame)
            faces_yn = faces_yn if faces_yn is not None else []
            
            for face in faces_yn:
                
                box = list(map(int, face[:4]))
                color = (255, 0, 0, 255)
                thickness = 2
                cv2.rectangle(cv2image, box, color, thickness, lineType=cv2.LINE_AA)
                try:
                    serial, confidence = face_recognizer.predict(gray_img[box[1]:box[1]+box[3], box[0]:box[0]+box[2]])
                    # print(serial, confidence)
                except:
                    serial, confidence = 0, 0
                position = (box[0], box[1] - 10)
                font = cv2.FONT_HERSHEY_SIMPLEX
                scale = 0.5
                thickness = 2
                color_red = (255, 0, 0, 255)
                color_green = (0, 255, 0, 255)
                if confidence > 70:
                    index = studentIDs.index(str(serial))
                    # print(index)
                    cv2.putText(cv2image, str(serial)+" "+"{:.2f}".format(confidence), position, font, scale, color_green, thickness, lineType=cv2.LINE_AA)
                    cv2.rectangle(cv2image, box, color_green, thickness, lineType=cv2.LINE_AA)
                    if str(serial) not in already_detected:
                        FaceRecognition.writeAttendance(file_path, index)
                        name = studentNames[index]
                        FaceRecognition.attendance_data.append([str(serial), name, datetime.datetime.now().strftime("%d/%m/%Y"), datetime.datetime.now().strftime("%H:%M:%S")])
                        already_detected.append(str(serial))
                else:
                    name = "Unknown"
                    cv2.putText(cv2image, name, position, font, scale, color_red, thickness, lineType=cv2.LINE_AA)
            
            img = Image.fromarray(cv2image)
            imgctk = ctk.CTkImage(light_image=img, dark_image=img, size=img.size)
            lmain.configure(image=imgctk)
            lmain.after(100, show_frame)
        show_frame()
        
    def create_attendance_frame(window, frame):
        # Create Attendance Frame
        attendance_frame = ctk.CTkFrame(master=frame)
        attendance_frame.grid(row=0, column=1, padx=20, pady=10)
        
        # Create Attendance Label
        attendance_label = ctk.CTkLabel(master=attendance_frame, text="Attendance", font=(font, 40, "bold"))
        attendance_label.pack()
        
        # Create Attendance Table
        tree_scrollbar = ttk.Scrollbar(attendance_frame, orient="vertical")
        tree_scrollbar.pack(side="right", fill="y")
        table = ttk.Treeview(attendance_frame, height=30, show="headings", yscrollcommand=tree_scrollbar.set)
        table.pack(fill="both", expand=True)
        tree_scrollbar.config(command=table.yview)
        
        # Configure Table
        style = ttk.Style()
        style.configure("Treeview.Heading", font=(font, 15, "bold"))
        style.configure("Treeview", font=(font, 15), rowheight=20)
        
        # Insert Table
        table['columns'] = ["ID", "Name", "Date", "Time"]
        table.column('#0', width=0, stretch='YES')
        table.column('ID', width=100, anchor='center')
        table.column('Name', width=200, anchor='center')
        table.column('Date', width=150, anchor='center')
        table.column('Time', width=150, anchor='center')
    
        table.heading('ID', text='ID', anchor='center')
        table.heading('Name', text='Name', anchor='center')
        table.heading('Date', text='Date', anchor='center')
        table.heading('Time', text='Time', anchor='center')
        
        for value in FaceRecognition.attendance_data:
            table.insert("", "end", values=value)
                
        # Back without save Button
        back_button = ctk.CTkButton(master=frame, fg_color='red',
                                    width=220, height=50,
                                    text="Back",
                                    font=(font, button_font_size, "bold"),
                                    command=lambda: FaceRecognition.back(window, frame))
        back_button.pack(side="bottom", pady=5)
         
        
    def __init__(self, window, file_path, studentIDs, studentNames, exam_time):
        # Create Master Frame
        master_frame = ctk.CTkFrame(master=window)
        master_frame.pack(fill="both", expand=True)
        master_frame.grid_rowconfigure(0, weight=1)
        master_frame.grid_columnconfigure(0, weight=1)
        master_frame.grid_columnconfigure(1, weight=1)
        
        for item in studentIDs:
            studentIDs[studentIDs.index(item)] = str(item)
        
        # Create Camera Frame
        FaceRecognition.create_camera_frame(master_frame, file_path, studentIDs, studentNames)
        FaceRecognition.create_attendance_frame(window, master_frame)

