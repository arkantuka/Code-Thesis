import customtkinter as ctk
import pages.main_page as mp
import pages.manage_student_data_page as mng_std
import openpyxl
import os
from tkinter import filedialog
from CTkMessagebox import CTkMessagebox

font = "THSarabunNew"
button_font_size = 30
header_font_size = 50

class StudentMenu:
    
    # File Dialog for file path
    def fileDialog(initial_directory):
        file_path = filedialog.askopenfilename(initialdir=initial_directory,
                                        title="Select A File",
                                        filetype=(("xlsx files", "*.xlsx"),("All Files", "*.*")))
        return file_path
                
    # Load Excel Data and save to new sheet
    def loadExcelData():
        try:
            parent_dir = "version.final/images"
            file_path = StudentMenu.fileDialog("/")
            workbook = openpyxl.load_workbook(file_path)
            sheets = workbook.sheetnames
            for sheet in sheets:
                new_workbook = openpyxl.Workbook()
                new_workbook.create_sheet(sheet)
                new_workbook.remove(new_workbook['Sheet'])
                for row in workbook[sheet].iter_rows(values_only=True):
                    # Strip \xa0 from text and add to New File
                    row = list(row)
                    row[0] = str(row[0])
                    row[0] = row[0].replace("\xa0", "")
                    new_workbook[sheet].append(row)
                new_workbook.save(f"version.final/students_data/{sheet}.xlsx")
                new_workbook.close()
                # Create Images Directory
                path = os.path.join(parent_dir, sheet)
                if not os.path.exists(path): os.mkdir(path)
                
            # Show Message
            CTkMessagebox(title="Information", 
                        message=f"File {file_path} loaded successfully.", 
                        icon="check", option_1="OK", option_2=None)
        except:
            pass
    
    # Open Excel Data and open new Page
    def openExcelData(window, frame):
        try:
            file_path = StudentMenu.fileDialog("version.final/students_data/")
            if file_path:
                frame.destroy()
                mng_std.ManageStudentPage(window, file_path)
        except:
            pass
        
    # Back Button Function
    def back(window, frame):
        frame.destroy()
        mp.MainPage(window)
        
    # Create Button
    def createButton(frame, label_name, command):
        button = ctk.CTkButton(master=frame, fg_color="gray",
                           text=label_name, height=50, width=230,
                           font=(font, button_font_size, "bold"),
                           command=command)
        return button
        
    def __init__(self, window):
        # Create Master Frame
        master_frame = ctk.CTkFrame(master=window)
        master_frame.pack(fill="both", expand=True)
        
        # Create Menu Label
        student_data_label = ctk.CTkLabel(master=master_frame,
                                  text="Manage Students Data Menu",
                                  font=(font, header_font_size, "bold"),
                                  padx=20, pady=10)
        student_data_label.pack(pady=(20, 0))
       
        # Create Buttons Frame
        buttons_frame = ctk.CTkFrame(master=master_frame)
        buttons_frame.place(relx=0.5, rely=0.5, anchor=ctk.CENTER)
        
        # Create Open Data Button
        open_data_button = StudentMenu.createButton(buttons_frame, 'Open Student Data',
                                                        lambda: StudentMenu.openExcelData(window, master_frame))
        open_data_button.grid(row=0, column=1, padx=20, pady=10)
        
        # Create Browse and Load File Button
        import_data_button = StudentMenu.createButton(buttons_frame, 'Import Excel File', 
                                                        lambda: StudentMenu.loadExcelData())
        import_data_button.grid(row=1, column=1, padx=20, pady=10)
        
        # Exit Button
        exit_button = ctk.CTkButton(master=master_frame, fg_color='red',
                                    width=200, height=50,
                                    text="Exit",
                                    font=(font, button_font_size, "bold"),
                                    command=window.destroy)
        exit_button.pack(side="bottom", pady=30)
        
        # Back Button
        back_button = ctk.CTkButton(master=master_frame, fg_color='green',
                                    width=200, height=50,
                                    text="Back",
                                    font=(font, button_font_size, "bold"),
                                    command=lambda: StudentMenu.back(window, master_frame))
        back_button.pack(side="bottom", pady=30)